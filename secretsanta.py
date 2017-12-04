"""
Clare Hanlon '18
November 24th, 2015
Script made in my own time for Kizuna secret santa event.

Modified 11/20/17 to work with xlsx spreadsheets
"""

from random import *
from smtplib import SMTP
from email.mime.text import MIMEText
from commands import *
from openpyxl import *
from unicodedata import normalize

input = "./santainfo17.xlsx"
output = "./santachain17.txt"

class Participant(object):

    def __init__(self, row):
        # info is a list of the information for each participant read from
        # the given document

        self.addr = row[5].value
        self.name = row[1].value
        self.wants = row[3].value
        self.notwants = row[4].value

    def __str__(self):
        result = ""
        result += "Name: " + self.name + "\n\n"
        result += "Wants: " + self.wants + "\n"
        result += "Does not want: " + self.notwants + "\n"
        return result

    def assign(self, bef, aft):
        # bef is which Participant is before this Participant in the list,
        # aft is who is after
        self.aft = aft
        self.bef = bef

def parseSheet(doc):
    wb = Workbook()
    wb = load_workbook(filename = doc)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    ppllist = []

    for i in range(2, ws.max_row+1):
        row = ws[i]
        # note: order of items within row depends on structure of survey

        # unicode/ASCII is some B U L L S H I T
        for cell in row:
            if type(cell.value) == unicode:
                cell.value = normalize('NFKD', cell.value).encode('ascii','ignore')
            elif cell.value is None:
                cell.value = "[Nothing written]"
            else: # so far the only other case i've encountered is float
                cell.value = str(cell.value)

        if row[2].value == "Yes":
            person = Participant(row)
            ppllist.append(person)

    return ppllist

def assignPeople(partlist):
    length = len(partlist)
    for i in range(length):
        bef = partlist[(i-1)%length]
        aft = partlist[(i+1)%length]
        part = partlist[i]
        part.assign(bef, aft)
        getoutput('echo "%s is receiving from %s." >> %s'%(part.name, bef.name, output))
        getoutput('echo "%s is giving to %s." >> %s'%(part.name, aft.name, output))

def emailtext(part):
    text = "Hello, %s!\n"%((part.name).rstrip())
    text += "Thank you for taking part in Kizuna's 2016 secret santa exchange. Below is some information about the person you will be giving a gift to.\n\n"
    text += str(part.aft)
    text += "The gift exchange will take place at 9:30 PM on Saturday, December 9th in the Danawell Multipurpose Room, so please get a gift for this person before then!\nAlso, please remember to write three hints as to who you are on your present. They don't have to be too specific/obvious, but they shouldn't be too general either (i.e. 'I am a Swarthmore student.').\n If for whatever reason you are not able to provide a gift in time for the exchange, please email chanlon1@swarthmore.edu as soon as possible so that we can remove you from the event. (Meaning if you don't give someone a gift, you won't get one, either!) If you show up to the gift exchange without a gift for your assigned person, the gift that was meant for you will be given to them.\n"
    text += "Thanks again for taking part in this event, and please feel free to reach out to your club leaders if you have any questions.\n"
    text += "\nSincerely," + "\n   Kizuna\n"
    text += "\nP.S. Remember to keep it secret!\n\n"
    return text

def email(part):
    name = part.name
    addr = part.addr
    text = emailtext(part)
    msg = MIMEText(text)
    msg['To'] = addr
    msg['From'] = "kizuna@sccs.swarthmore.edu"
    msg['Subject'] = "Kizuna Secret Santa Assignment"

    s = SMTP()
    s.connect("allspice.cs.swarthmore.edu")
    try:
        s.sendmail(msg['From'], [msg['To']], msg.as_string())
    except Exception, e:
        print "Failed to email %s"%(addr)
        return None
        s.quit()
    s.close()
    return

def main():
    getoutput("rm -f ./santachain17.txt")
    doc = input
    #do things
    ppl = parseSheet(doc)
    shuffle(ppl)
    #assign
    assignPeople(ppl)
    #for each item in list:
    for person in ppl:
        #email
        email(person)

main()
